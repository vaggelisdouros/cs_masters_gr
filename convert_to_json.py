import openpyxl
import json

# Το όνομα του αρχείου Excel σου
excel_file = 'cs_masters_study_areas.xlsx'

# Ο αριθμός της στήλης που περιέχει το όνομα του τμήματος
department_col = 1 # Π.χ. στήλη A

# Ο αριθμός της στήλης που περιέχει το όνομα του πανεπιστημίου
university_col = 2 # Π.χ. στήλη B

# Ο αριθμός της στήλης που περιέχει την πόλη
city_col = 3 # Π.χ. στήλη C

# Ο αριθμός της στήλης όπου ξεκινάει το πρώτο μεταπτυχιακό πρόγραμμα
# Θα πρέπει να το αλλάξεις ανάλογα με το αρχείο σου.
first_program_col = 4 # Π.χ. στήλη D

# Οι αριθμοί των στηλών που περιέχουν τους τομείς σπουδών
first_study_area_col = 6  # Στήλη F
last_study_area_col = 18 # Στήλη Q

# Φόρτωμα του αρχείου Excel
workbook = openpyxl.load_workbook(excel_file)
sheet = workbook.active

# Συνάρτηση για να μετατρέπει μια συμβολοσειρά με κόμματα σε λίστα
def get_list_from_cell(cell_value):
    if isinstance(cell_value, str):
        # 1. Αντικαθιστούμε τις αλλαγές γραμμής με ένα κενό
        single_line_text = cell_value.replace('\n', ' ')
        
        # 2. Χωρίζουμε τη συμβολοσειρά με κόμμα και αφαιρούμε τα κενά από κάθε στοιχείο
        # Χρησιμοποιεί το "if item.strip()" για να αποφύγει κενά στοιχεία στη λίστα
        return [item.strip() for item in single_line_text.split(',') if item.strip()]
    elif cell_value:
        return [cell_value]
    return []

programs_list = []
# Ξεκινάμε από τη δεύτερη γραμμή, υποθέτοντας ότι η πρώτη είναι οι επικεφαλίδες
for row_num in range(2, sheet.max_row + 1):
    # Παίρνουμε τις βασικές πληροφορίες του τμήματος
    # Χρησιμοποιούμε τη νέα συνάρτηση για να δημιουργήσουμε λίστες
    department_names = get_list_from_cell(sheet.cell(row=row_num, column=department_col).value)
    university_names = get_list_from_cell(sheet.cell(row=row_num, column=university_col).value)
    cities = get_list_from_cell(sheet.cell(row=row_num, column=city_col).value)

    # Αν δεν υπάρχει όνομα τμήματος, παραλείπουμε τη γραμμή
    if not department_names:
        continue

    # Προσθήκη: Δημιουργία της λίστας με τους τομείς σπουδών
    study_areas = []
    # Παίρνουμε τις επικεφαλίδες των στηλών (η πρώτη γραμμή)
    header_row = [cell.value for cell in sheet[1]]
    for col_num_study in range(first_study_area_col, last_study_area_col + 1):
        cell_value = sheet.cell(row=row_num, column=col_num_study).value
        # Ελέγχουμε αν η τιμή είναι 'YES' (αγνοώντας τα κενά και την πεζή/κεφαλαία γραφή)
        if isinstance(cell_value, str) and cell_value.strip().upper() == 'YES':
            study_areas.append(header_row[col_num_study - 1])
    
    # Διατρέχουμε τις στήλες ανά δύο για τα μεταπτυχιακά και τα δίδακτρα
    # Ο βρόχος τώρα σταματάει πριν τους τομείς σπουδών.
    for col_num in range(first_program_col, first_study_area_col, 2):
        
        program_cell = sheet.cell(row=row_num, column=col_num)
        tuition_cell = sheet.cell(row=row_num, column=col_num + 1)

        # Ελέγχουμε αν υπάρχει όνομα μεταπτυχιακού για να συνεχίσουμε
        if program_cell.value:
            program_name = program_cell.value
            tuition = tuition_cell.value if tuition_cell.value else None
            link = program_cell.hyperlink.target if program_cell.hyperlink else ''
            
            program_entry = {
                'department_name': department_names,
                'university_name': university_names,
                'city': cities,
                'program_name': program_name,
                'link': link,
                'tuition': tuition,
                'study_area': study_areas # Προσθήκη του νέου πεδίου
            }
            programs_list.append(program_entry)

# Αποθήκευση των δεδομένων σε αρχείο JSON
json_file = 'data.json'
with open(json_file, 'w', encoding='utf-8') as f:
    json.dump(programs_list, f, ensure_ascii=False, indent=4)

print(f"Το αρχείο '{excel_file}' μετατράπηκε με επιτυχία σε '{json_file}'.")